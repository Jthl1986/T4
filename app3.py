import pandas as pd
import streamlit as st
import lxml
import urllib.request
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from streamlit_lottie import st_lottie
import requests
import streamlit.components.v1 as components
import json
import openpyxl
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
import time

st.set_page_config(page_title="AgroAppCredicoop",page_icon="🌱",layout="wide") 

@st.experimental_memo
def load_unpkg(src: str) -> str:
    return requests.get(src).text


HTML_2_CANVAS = load_unpkg("https://unpkg.com/html2canvas@1.4.1/dist/html2canvas.js")
JSPDF = load_unpkg("https://unpkg.com/jspdf@latest/dist/jspdf.umd.min.js")
BUTTON_TEXT = "Create PDF"

def copy_button():
    copy_button = Button(label="Copiar tabla")
    copy_button.js_on_event("button_click", CustomJS(args=dict(df=st.session_state.dfa.to_csv(sep='\t')), code="""
        navigator.clipboard.writeText(df);
        """))
    no_event = streamlit_bokeh_events(
        copy_button,
        events="GET_TEXT",
        key="get_text",
        refresh_on_update=True,
        override_height=75,
        debounce_time=0)
    
def css():
    # CSS to inject contained in a string
    hide_table_row_index = """
            <style>
            thead tr th:first-child {display:none}
            tbody th {display:none}
            </style>
            """
    # Inject CSS with Markdown
    st.markdown(hide_table_row_index, unsafe_allow_html=True)

def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

def app():
    st.title("🐮 Valuación de hacienda")
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de hacienda: ', ["Ternero             ", "Novillito       ", "Ternera             ", "Vaquillona        ", "Vaca                "])
    cantidad = form.number_input("Ingrese cantidad de cabezas: ", step=1)
    peso = form.number_input("Ingrese peso: ", step=1)
    submit = form.form_submit_button("Ingresar")
    df=pd.read_html('https://www.monasterio-tattersall.com/precios-hacienda') #leo la tabla de la página
    hacienda = df[0] 
    categoria = hacienda.Categoría 
    promedio = hacienda.Promedio
    tabla = pd.DataFrame({'categoria':categoria,'promedio':promedio}) #creo un dataframe con categoria y promedio
    ternero=tabla[0:4] 
    novillito=tabla[4:7]
    ternera=tabla[7:11]
    vaquillona=tabla[11:14]
    vaca=tabla[19:20]
    fecha=(tabla[25:26].values)[0][0]
    ternero160=int(ternero.promedio[0][2:5])
    ternero180=int(ternero.promedio[1][2:5])
    ternero200=int(ternero.promedio[2][2:5])
    ternero230=int(ternero.promedio[3][2:5])
    novillo260=int(novillito.promedio[4][2:5])
    novillo300=int(novillito.promedio[5][2:5])
    novillo301=int(novillito.promedio[6][2:5])
    ternera150=int(ternera.promedio[7][2:5])
    ternera170=int(ternera.promedio[8][2:5])
    ternera190=int(ternera.promedio[9][2:5])
    ternera210=int(ternera.promedio[10][2:5])
    vaquillona250=int(vaquillona.promedio[11][2:5])
    vaquillona290=int(vaquillona.promedio[12][2:5])
    vaquillona291=int(vaquillona.promedio[13][2:5])
    vacas=int(vaca.promedio[19][2:8])
    def constructor():
        def valores():
            if tipo == 'Ternero             ' and peso < 160:
                valor = ternero160*cantidad*peso
            elif tipo == 'Ternero             ' and peso < 180:
                valor = ternero180*cantidad*peso
            elif tipo == 'Ternero             ' and peso <= 200:
                valor = ternero200*cantidad*peso
            elif tipo == 'Ternero             ' and peso > 200:
                valor = ternero230*cantidad*peso
            elif tipo == 'Ternero             ' and peso == 0:
                valor = ternero200*cantidad*peso
            elif tipo == 'Novillito       ' and peso < 260:
                valor = novillo260*cantidad*peso
            elif tipo == 'Novillito       ' and peso <= 300:
                valor = novillo300*cantidad*peso
            elif tipo == 'Novillito       ' and peso > 300:
                valor = novillo301*cantidad*peso
            elif tipo == 'Novillito       ' and peso == 0:
                valor = novillo300*cantidad*peso
            elif tipo == 'Ternera             ' and peso < 150:
                valor = ternera150*cantidad*peso
            elif tipo == 'Ternera             ' and peso < 170:
                valor = ternera170*cantidad*peso
            elif tipo == 'Ternera             ' and peso <= 190:
                valor = ternera190*cantidad*peso
            elif tipo == 'Ternera             ' and peso > 190:
                valor = ternera210*cantidad*peso
            elif tipo == 'Ternera             ' and peso == 0:
                valor = ternera190*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso < 250:
                valor = vaquillona250*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso <= 290:
                valor = vaquillona290*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso > 290:
                valor = vaquillona291*cantidad*peso
            elif tipo == 'Vaquillona        ' and peso == 0:
                valor = vaquillona290*cantidad*peso
            elif tipo == 'Vaca                ':
                valor = vacas*cantidad
            valor = int(valor*0.9)
            return valor #valor de ajuste
        valor=valores()
        d = [tipo, cantidad, peso, valor]
        return d
    metalista=[]
    if "dfa" not in st.session_state:
        st.session_state.dfa = pd.DataFrame(columns=("Categoría", "Cantidad", "Peso", "Valuación"))
    if submit:
        metalista.append(constructor())
        dfb = pd.DataFrame(metalista, columns=("Categoría", "Cantidad", "Peso", "Valuación"))
        st.session_state.dfa = pd.concat([st.session_state.dfa, dfb])
    css()
    valuacion_total = st.session_state.dfa['Valuación'].sum()
    right.metric('La valuación total de hacienda es: ', '${:,}'.format(valuacion_total))

    del_button = right.button("Borrar última fila")
    if del_button and len(st.session_state.dfa) > 0:
        st.session_state.dfa = st.session_state.dfa.iloc[:-1]

    right.write("Tabla para copiar:")
    right.table(st.session_state.dfa.style.format({"Cantidad":"{:.0f}", "Peso":"{:.0f}", "Valuación":"${:,}"}))
    right.write(f'Los precios considerados son de la {fecha}')
    promedios = pd.DataFrame(
        {'Categoria': ['Ternero', 'Novillo', 'Ternera', 'Vaquillonas'],
         'Peso': ['180', '260', '170','250']})
    st.write(f'Pesos promedio para tipo de hacienda (en caso que no se informe el peso). En vacas poner peso cero')
    st.table(promedios.assign(hack='').set_index('hack'))
    
def app1():
    df2=pd.read_html('https://www.cotagroweb.com.ar/pizarra/')
    data2 = df2[0]
    # psoja= 58760 en caso que falle precio cotagro habilitar esta línea
    psoja = data2.iloc[0,1]
    ppsoja = int(psoja[1:])
    pmaiz= data2.iloc[1,1]
    ppmaiz = int(pmaiz[1:])
    ptrigo= data2.iloc[2,1]
    pptrigo = int(ptrigo[1:])
    pgira= data2.iloc[4,1]
    ppgira = 74400 #int(pgira[1:])
    ppsorgo = 44400
    fecha = data2.columns[1][7:]
    st.title("🌾 Valuación de granos")
    st.write(f'Precios de pizarra del Mercado de Rosario al {fecha}')
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Soja", '${:,}'.format(int(ppsoja)))
    col2.metric("Trigo", '${:,}'.format(int(pptrigo)))
    col3.metric("Maíz", '${:,}'.format(int(ppmaiz)))
    col4.metric("Sorgo", '${:,}'.format(int(ppsorgo)))
    col5.metric("Girasol",'${:,}'.format(int(ppgira)))
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de grano: ', ["Soja","Trigo","Maíz","Sorgo","Girasol"])
    cantidad = form.number_input("Ingrese toneladas: ", step=1)
    submit = form.form_submit_button("Ingresar")
    def lista():
        def valor():
            if tipo == "Soja":
                precio = ppsoja
            elif tipo == "Trigo":
                precio = pptrigo
            elif tipo == "Maíz":
                precio = ppmaiz
            elif tipo == "Sorgo":
                precio = ppsorgo
            else:
                precio = ppgira
            return int(cantidad*precio)
        valor = valor()
        lista = [tipo, cantidad, valor]
        return lista
    cereales=[]
    if "dfs" not in st.session_state:
        st.session_state.dfs = pd.DataFrame(columns=("Tipo grano", "Cantidad (tn)", "Valuación"))
    if submit:
        cereales.append(lista())
        dfd = pd.DataFrame(cereales, columns=("Tipo grano", "Cantidad (tn)", "Valuación"))
        st.session_state.dfs = pd.concat([st.session_state.dfs, dfd])
    css()
    valuacion_total = st.session_state.dfs['Valuación'].sum()
    right.metric('La valuación total de granos es: ', '${:,}'.format(valuacion_total))
    del_button = right.button("Borrar última fila")
    if del_button and len(st.session_state.dfs) > 0:
        st.session_state.dfs = st.session_state.dfs.iloc[:-1]
    right.write("Tabla para copiar:")
    right.table(st.session_state.dfs.style.format({"Cantidad (tn)":"{:.0f}", "Valuación":"${:,}"}))

def app2():
    if "ingresos_totales" not in st.session_state:
        st.session_state["ingresos_totales"] = 0
    st.title("🚜 Servicios agrícolas")
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form")
    tipo = form.selectbox('Ingrese tipo de servicio: ', ["Cosecha","Siembra","Pulverización","Laboreos"])
    cantidad = form.number_input("Ingrese superficie (has): ", step=1)
    precio = form.number_input("Ingrese precio por ha", step=1)
    submit = form.form_submit_button("Ingresar")
    valorminc = 9000 #valor minimo cosecha
    valormaxc = 16000 #valor maximo cosecha
    valors = 7500 #valor referencia siembra
    valormins = valors*0.50 #valor minimo siembra
    valormaxs = valors*1.50 #valor maximo siembra
    
    def lista():
        def valor():
            return cantidad*precio
        valor = valor()
        lista = [tipo, cantidad, precio, valor]
        return lista
    servagro=[]
    if "dfx" not in st.session_state:
        st.session_state.dfx = pd.DataFrame(columns=("Categoría", "Superficie(ha)", "Precio", "Ingreso estimado"))
    if submit:
        servagro.append(lista())
        st.session_state["ingresos_totales"] += cantidad*precio
        dfy = pd.DataFrame(servagro, columns=("Categoría", "Superficie(ha)", "Precio", "Ingreso estimado"))
        st.session_state.dfx = pd.concat([st.session_state.dfx, dfy])
        if tipo == 'Cosecha' and (precio > valormaxc or precio < valorminc):
            st.warning("ALERTA! El precio por ha de cosecha cargado es fuera de los promedios de mercado. Ver precios de referencia abajo")
        elif tipo == 'Siembra' and (precio > valormaxs or precio < valormins):
            st.warning("ALERTA! El precio por ha de siembra cargado es fuera de los promedios de mercado. Ver precios de referencia abajo")
        else:
            pass
    
    right.metric('Los ingresos totales por servicios agrícolas son: ', "${:,}".format(st.session_state["ingresos_totales"]))    

    delete_last_row = right.button("Borrar última fila")
    if delete_last_row:
        if not st.session_state.dfx.empty:
            st.session_state["ingresos_totales"] -= st.session_state.dfx["Ingreso estimado"].iloc[-1]
            st.session_state.dfx = st.session_state.dfx.iloc[:-1]
    css()
    
    right.write("Tabla para copiar:")
    right.table(st.session_state.dfx.style.format({"Superficie(ha)":"{:.0f}", "Precio":"${:,}", "Ingreso estimado":"${:,}"}))
    
    
    def mostrar_precios_referencia(tipo_servicio, imagen):
        expander = st.expander(f"Ver precios de referencia - {tipo_servicio}")
        expander.image(imagen)
    mostrar_precios_referencia("Cosecha Soja", "https://www.agrocontratistas.com.ar/img/Precios/SOJA202303.jpg")
    mostrar_precios_referencia("Cosecha Maíz", "https://www.agrocontratistas.com.ar/img/Precios/MAIZ202303.jpg")
    mostrar_precios_referencia("Siembra y Laboreos", "https://www.agrocontratistas.com.ar/img/Precios/Labores_202302.jpg")
    return st.session_state.dfx
    
def app3():
    st.title("⛅️ Estado de los campos")
    with st.expander("Recomendaciones de interpretación"):
     st.write("""
         - Para ver el panorama general de sequía ir a ¿Qué zonas estan en sequía? y buscar en "unidad administrativa de nivel 2" la localidad donde estan los campos
         - En caso de estar en área de sequía ver la sección "Evolución de sequías entre dos períodos" para ver si se registraron mejoras en los ultimos meses.
         - En la sección ¿Hace cuanto que no llueve? se puede ver la última información de precipitaciones
         - Tener en cuenta que el mapa de calor se conforma con la información recolectada de las estaciones por lo que algunas áreas con pocas estaciones (como por ejemplo zona centro este de Santa Fe) pueden verse influenciadas por estaciones más lejanas
     """)
    components.iframe("https://dashboard.crc-sas.org/informes/como-estamos/", height = 1500)
    st.caption("Datos extraidos de https://sissa.crc-sas.org/novedades/publicaciones-y-reportes-tecnicos/")
    

def app4():
    st.title("🌽 Planteo productivo")
    region = st.selectbox('Región: ', ["N Bs As/S Sta Fe","Oeste Bs As", "SO Bs As", "SE Bs As", "S Cordoba", "S Entre Ríos","Salta", "S del Estero"])
    left, right = st.columns(2)
    left.write("Completar:")
    form = left.form("template_form") 
    tipo = form.selectbox('Tipo de cultivo: ', ["Soja 1ra", "Soja 2da", "Trigo","Maíz","Girasol", "Sorgo", "Cebada Forrajera", "Cebada Cervecera"])
    propio = form.selectbox('Campos: ', ["Propios","Arrendados","Aparcería"])
    cantidad = form.number_input("Superficie (has): ", step=1)
    rinde = form.number_input("Rendimiento informado (en tn)")
    submit = form.form_submit_button("Ingresar")
        
    # API tipo de cambio
    url = "https://www.dolarsi.com/api/api.php?type=valoresprincipales"
    response = requests.get(url)
    if response.status_code == 200:
        api_data = response.json()
        value = api_data[0]['casa']['venta']
        value2 = value.replace(',', '.')
        dol = float(value2)
    else:
        print("Failed to retrieve data")
        
    right.metric("Dolar oficial", '${:,}'.format(float(dol)))
    right.write("Cuadro gastos:")
    form2 = right.form("template_form2") 
    gast = form2.number_input("Gastos de estructura", step=1)
    arrendamiento = form2.number_input("Gastos de arrendamiento", step=1)
    aparceria = form2.number_input("Porcentaje de aparcería", step=1)
    submit2 = form2.form_submit_button("Ingresar")
    
    #unpacking
    url = 'https://raw.githubusercontent.com/Jthl1986/T4/master/dataframe.xlsx'
    r = requests.get(url)
    
    if r.status_code == 200:
        with open('temp.xlsx', 'wb') as f:
            f.write(r.content)
    
        workbook = openpyxl.load_workbook('temp.xlsx')
        worksheet = workbook.active
    
        header = [cell.value for cell in next(worksheet.iter_rows())]
        data = [cell.value for row in worksheet.iter_rows(min_row=2) for cell in row]
    
        result = dict(zip(header, data))
    
        for key, value in result.items():
            globals()[key] = value
    
    else:
        print("No se pudo descargar el archivo")
    
    #Matriz de producción
    regiones = ["N Bs As/S Sta Fe","Oeste Bs As", "SO Bs As", "SE Bs As", "S Cordoba", "S Entre Ríos","Salta", "S del Estero"]
    cultivos = ["Soja 1ra", "Soja 2da", "Trigo","Maíz","Girasol", "Sorgo", "Cebada Forrajera", "Cebada Cervecera"]
    
    # Precios por región y tipo de cultivo
    precios = [
        [sojaprice1, soja2price1, trigoprice1, maizprice1, giraprice2, sorgoprice1, cebadaprice2, cebadaprice1],  # N Bs As/S Sta Fe
        [sojaprice2, soja2price2, trigoprice3, maizprice1, giraprice2, sorgoprice1, cebadaprice2, cebadaprice1],   # Oeste Bs As
        [sojaprice2, soja2price2, trigoprice3, maizprice2, giraprice1, sorgoprice1, cebadaprice2, cebadaprice1],   # SO Bs As
        [sojaprice3, soja2price2, trigoprice2, maizprice2, giraprice1, sorgoprice1, cebadaprice2, cebadaprice1], # SE Bs As
        [sojaprice1, soja2price1, trigoprice1, maizprice1, giraprice3, sorgoprice1, cebadaprice2, cebadaprice1],  # S Cordoba
        [sojaprice1, soja2price1, trigoprice1, maizprice1, giraprice3, sorgoprice1, cebadaprice2, cebadaprice1],   # S Entre Ríos
        [sojaprice1, soja2price1, trigoprice1, maizprice1, giraprice2, sorgoprice1, cebadaprice2, cebadaprice1],    # Salta
        [sojaprice1, soja2price1, trigoprice1, maizprice1, giraprice2, sorgoprice1, cebadaprice2, cebadaprice1]    # S del Estero
    ]
    
    # Costos por región y tipo de cultivo
    costos = [
        [sojacost5, soja2cost3, trigocost1, maizcost1, giracost3, sorgocost3, cebadacost2, cebadacost1],    # N Bs As/S Sta Fe
        [sojacost8, soja2cost4, trigocost7, maizcost5, giracost2, sorgocost1, cebadacost2, cebadacost1],     # Oeste Bs As
        [sojacost3, soja2cost4, trigocost5, maizcost4, giracost1, sorgocost1, cebadacost2, cebadacost1],     # SO Bs As
        [sojacost4, soja2cost4, trigocost3, maizcost3, giracost1, sorgocost1, cebadacost2, cebadacost1],    # SE Bs As
        [sojacost7, soja2cost1, trigocost6, maizcost6, giracost3, sorgocost4, cebadacost2, cebadacost1],    # S Cordoba
        [sojacost6, soja2cost2, trigocost2, maizcost2, giracost3, sorgocost2, cebadacost2, cebadacost1],     # S Entre Ríos
        [sojacost2, soja2cost3, trigocost4, maizcost1, giracost3, sorgocost2, cebadacost2, cebadacost1],     # Salta
        [sojacost1, soja2cost3, trigocost4, maizcost1, giracost3, sorgocost2, cebadacost2, cebadacost1]      # S del Estero
    ]
    
    gastos = [
        [sojagc5, soja2gc3, trigogc1, maizgc1, giragc3, sorgogc3, cebadagc2, cebadagc1],    # N Bs As/S Sta Fe
        [sojagc8, soja2gc4, trigogc7, maizgc4, giragc2, sorgogc1, cebadagc2, cebadagc1],     # Oeste Bs As
        [sojagc3, soja2gc4, trigogc5, maizgc3, giragc1, sorgogc1, cebadagc2, cebadagc1],     # SO Bs As
        [sojagc4, soja2gc4, trigogc3, maizgc2, giragc1, sorgogc1, cebadagc2, cebadagc1],    # SE Bs As
        [sojagc7, soja2gc1, trigogc6, maizgc5, giragc3, sorgogc4, cebadagc2, cebadagc1],    # S Cordoba
        [sojagc6, soja2gc2, trigogc2, maizgc1, giragc3, sorgogc2, cebadagc2, cebadagc1],     # S Entre Ríos
        [sojagc2, soja2gc3, trigogc4, maizgc1, giragc3, sorgogc2, cebadagc2, cebadagc1],     # Salta
        [sojagc1, soja2gc3, trigogc4, maizgc1, giragc3, sorgogc2, cebadagc2, cebadagc1]      # S del Estero
    ]
    
    precio = precios[regiones.index(region)][cultivos.index(tipo)]
    costo = costos[regiones.index(region)][cultivos.index(tipo)]
    gasto = gastos[regiones.index(region)][cultivos.index(tipo)]
    
    # Imprimir la lista de datos        
    def lista():
        def valor1():
            return precio*dol*rinde*cantidad
        valors = round(valor1())
        
        def costo1():
            return costo*dol*cantidad
        cost = round(costo1())
        
        def gc1():
            return gasto*valors
        gc = round(gc1())
        
        def neto():
            return valors-cost-gc
        net = round(neto())
        
        lista = [region, propio, tipo, cantidad, rinde, valors, cost, gc, net]
        return lista
    datos = []
    if "dfp" not in st.session_state:
        st.session_state.dfp = pd.DataFrame(columns=('Región', 'Campos', 'Cultivo', 'Superficie (has)', 'Rinde', 'Ingreso', 'Costos directos', 'Gastos comercialización','Margen bruto'))
    if submit:
        datos.append(lista())
        dfo = pd.DataFrame(datos, columns=('Región', 'Campos','Cultivo', 'Superficie (has)', 'Rinde', 'Ingreso', 'Costos directos','Gastos comercialización', 'Margen bruto'))
        st.session_state.dfp = pd.concat([st.session_state.dfp, dfo])
    st.dataframe(st.session_state.dfp.style.format({"Superficie (has)":"{:.0f}", "Rinde":"{:,}", "Ingreso":"${:,}", "Costos directos":"${:,}", "Gastos comercialización":"${:,}", "Margen bruto":"${:,}"}))
    css()
    delete_last_row = left.button("Borrar última fila")
    if delete_last_row:
        if not st.session_state.dfp.empty:
            st.session_state["ingresos_totales"] -= st.session_state.dfp["Ingreso"].iloc[-1]
            st.session_state.dfp = st.session_state.dfp.iloc[:-1]
    if submit2:
        st.session_state.df1 = [arrendamiento, gast, aparceria]
        
    

def app5():
    st.header("Cuadro resumen")
    left, right = st.columns(2)
    css()
   
    # Obtener los dataframes existentes o None si no existen
    dfp = getattr(st.session_state, 'dfp', None)
    dfs = getattr(st.session_state, 'dfs', None)
    dfx = getattr(st.session_state, 'dfx', None)
    dfa = getattr(st.session_state, 'dfa', None)
    df1 = getattr(st.session_state, 'df1', None)
    
   
    if dfp is not None:
        st.subheader("Planteo productivo")
        ingtotal = st.session_state.dfp['Ingreso'].sum()
        costtotal = st.session_state.dfp['Costos directos'].sum()
        gctotal = st.session_state.dfp['Gastos comercialización'].sum()
        mbtotal = st.session_state.dfp['Margen bruto'].sum()
    if df1 is not None:
        left, middle, right = st.columns(3)
        arrend = st.session_state.df1[0]
        gas = st.session_state.df1[1]
        result = int(mbtotal)-int(arrend)-int(gas)
        # Crear una lista de diccionarios con los datos
        
        # Define el estilo CSS
        csss = """
        <style>
            .custom-table tbody tr:last-child td {
                font-weight: bold;
            }
        """

        data = [
            {'Concepto': 'Facturación campaña', 'Total': '${:,}'.format(round(ingtotal))},
            {'Concepto': 'Costos directos', 'Total': '${:,}'.format(round(costtotal))},
            {'Concepto': 'Gastos comercialización', 'Total': '${:,}'.format(round(gctotal))},
            {'Concepto': 'Margen bruto total', 'Total': '${:,}'.format(round(mbtotal))},
            {'Concepto': 'Arrendamiento', 'Total': '${:,}'.format(arrend)},
            {'Concepto': 'Gastos estructura', 'Total': '${:,}'.format(gas)},
            {'Concepto': 'Generación operativa de fondos', 'Total': '${:,}'.format(result)}
            ]
        # Apply CSS styles to the table
        left.markdown(csss, unsafe_allow_html=True)
        left.markdown('<table class="custom-table">{}</table>'.format(pd.DataFrame(data).to_html(index=False, classes="custom-table", header=False)), unsafe_allow_html=True)
        st.write('\n\n')
        st.write('\n\n')
        st.dataframe(dfp.style.format({"Superficie (has)":"{:.0f}", "Rinde":"{:,}", "Ingreso":"${:,}", "Costos directos":"${:,}", "Gastos comercialización":"${:,}", "Margen bruto":"${:,}"})) 
        
        # Agrupar datos por tipo de cultivo y sumar las hectáreas correspondientes
        chart = st.session_state.dfp.groupby('Cultivo')['Superficie (has)'].sum()
        
        # Crear gráfico de barras con los datos obtenidos
        middle.bar_chart(chart)
        
    if dfp is not None and df1 is None:
        st.write ("Sin planteo productivo o falta cargar gastos de estructura")
        
    if dfs is not None:
        left, right = st.columns(2)
        left.subheader("Existencias de granos")
        left.table(dfs.style.format({"Cantidad (tn)":"{:.0f}", "Valuación":"${:,}"}))
       
    if dfx is not None:
        left, right = st.columns(2)
        left.subheader("Ingresos Servicios agrícolas")
        left.table(dfx.style.format({"Superficie(ha)":"{:.0f}", "Precio":"${:,}", "Ingreso estimado":"${:,}"}))
   
    if dfa is not None:
        left, right = st.columns(2) 
        left.subheader("Existencias de hacienda")
        left.table(dfa.style.format({"Cantidad":"{:.0f}", "Peso":"{:.0f}", "Valuación":"${:,}"}))
        
        
    #topLeftMargin * 20 es donde manejas el ancho
    #allowTaint: true, scale: 3  es la definicion
    if st.button(BUTTON_TEXT):
        components.html(
                f"""
        <script>{HTML_2_CANVAS}</script>
        <script>{JSPDF}</script>
        <script>
        const html2canvas = window.html2canvas
        const {{ jsPDF }} = window.jspdf
        
        const streamlitDoc = window.parent.document;
        const stApp = streamlitDoc.querySelector('.main > .block-container');
        
        const buttons = Array.from(streamlitDoc.querySelectorAll('.stButton > button'));
        const pdfButton = buttons.find(el => el.innerText === '{BUTTON_TEXT}');
        const docHeight = stApp.scrollHeight;
        const docWidth = stApp.scrollWidth;
        
        let topLeftMargin = 30;
        let pdfWidth = docHeight + (topLeftMargin * 17);
        let pdfHeight = (pdfWidth * 1.5) + (topLeftMargin * 2);
        let canvasImageWidth = docWidth;
        let canvasImageHeight = docHeight;
        
        let totalPDFPages = Math.ceil(docHeight / pdfHeight)-1;
        
        pdfButton.innerText = 'Creating PDF...';
        
        html2canvas(stApp, {{ allowTaint: true, scale: 3 }}).then(function (canvas) {{
        
            canvas.getContext('2d');
            let imgData = canvas.toDataURL("image/jpeg", 1.0);
        
            let pdf = new jsPDF('p', 'px', [pdfWidth, pdfHeight]);
            pdf.addImage(imgData, 'JPG', topLeftMargin, topLeftMargin, canvasImageWidth, canvasImageHeight);
        
            for (var i = 1; i <= totalPDFPages; i++) {{
                pdf.addPage();
                pdf.addImage(imgData, 'JPG', topLeftMargin, -(pdfHeight * i) + (topLeftMargin*4), canvasImageWidth, canvasImageHeight);
            }}
        
            pdf.save('test.pdf');
            pdfButton.innerText = '{BUTTON_TEXT}';
        }})
        </script>
        """,
                    height=0,
                    width=0,
                )

        
#configuraciones de página   
lottie_book = load_lottieurl('https://assets7.lottiefiles.com/packages/lf20_d7OjnJ.json')
with st.sidebar:
    st.title('Agro App')
    st.markdown("---")
my_button = st.sidebar.radio("Modulos",('Planteo productivo', 'Condiciones climáticas', 'Tenencia granos', 'Tenencia hacienda', 'Servicios agrícolas', 'Cuadro resumen'))
if my_button == 'Tenencia hacienda':
    app()
elif my_button == 'Tenencia granos':
    app1()
elif my_button == 'Servicios agrícolas':
    app2()
elif my_button == 'Condiciones climáticas':
    app3()
elif my_button == 'Cuadro resumen':
    app5()
else:    
    app4()
with st.sidebar:
    st.markdown("---")
    st_lottie(lottie_book, speed=0.5, height=50, key="initial")
    st.markdown("---")
    st.caption("Desarrollado por JSantacecilia para Equipo Agro Banco Credicoop")
    
    
# Mantenimiento app
# psorgo l149
# parametros servicios agricolas l210 - 211 -212 y expanders l256, 258 y 260

#hacienda st.session_state.dfa
#granos st.session_state.dfs
#servicios st.session_state.dfx

